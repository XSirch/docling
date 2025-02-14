import streamlit as st
import sqlite3
from bcrypt import hashpw, gensalt, checkpw
import os
import re
import openpyxl
from docling.document_converter import DocumentConverter, PdfFormatOption, WordFormatOption, ExcelFormatOption
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from imagepdf.imagepdf import extract_images_from_pdf, get_image_description

# Configuração de pastas
UPLOAD_FOLDER = "uploads"
CONVERTED_FOLDER = "converted"
IMAGES_FOLDER = "images"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CONVERTED_FOLDER, exist_ok=True)
os.makedirs(IMAGES_FOLDER, exist_ok=True)

# Configuração do banco de dados
DB_NAME = "users.db"

def init_db():
    """Inicializa o banco de dados com a tabela de usuários."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

def add_user(username: str, password: str):
    """Adiciona um novo usuário ao banco de dados."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    hashed_password = hashpw(password.encode(), gensalt())
    try:
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hashed_password))
        conn.commit()
        st.success(f"Usuário '{username}' adicionado com sucesso!")
    except sqlite3.IntegrityError:
        st.error(f"O usuário '{username}' já existe.")
    conn.close()

def authenticate(username: str, password: str) -> bool:
    """Autentica o usuário comparando com as credenciais no banco de dados."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT password FROM users WHERE username = ?", (username,))
    result = cursor.fetchone()
    conn.close()
    if result and checkpw(password.encode(), result[0]):
        return True
    return False

# Configuração do Docling para PDF, DOCX e XLSX
def create_converter():
    pdf_pipeline_options = PdfPipelineOptions(
        do_ocr=True,
        do_table_structure=True,
        extract_images=True,
        images_output_path=IMAGES_FOLDER,
        table_structure_options={"mode": "accurate"}
    )
    pdf_format_option = PdfFormatOption(pipeline_options=pdf_pipeline_options)

    word_format_option = WordFormatOption()
    excel_format_option = ExcelFormatOption()

    return DocumentConverter(
        format_options={
            InputFormat.PDF: pdf_format_option,
            InputFormat.DOCX: word_format_option,
            InputFormat.XLSX: excel_format_option
        }
    )

converter = create_converter()

# Função para processar arquivos PDF (com imagens)
def process_pdf_with_images_and_text(pdf_path, output_folder):
    """Processa o PDF para incluir texto, tabelas e imagens com descrições logo após cada <!-- image -->"""
    image_paths = extract_images_from_pdf(pdf_path, output_folder)
    
    # Extrai o conteúdo textual do PDF
    result = converter.convert(pdf_path)
    markdown_content = result.document.export_to_markdown()

    # Substituir `<!-- image -->` pela descrição correspondente
    image_index = 0
    def replace_placeholder(match):
        """Substitui <!-- image --> pela imagem e descrição correspondente."""
        nonlocal image_index
        if image_index < len(image_paths):
            image_path = image_paths[image_index]
            description = get_image_description(image_path)
            markdown_image = f"![{description}]({image_path})\n"
            image_index += 1
            return f"<!-- image -->\n{markdown_image}"
        return match.group(0)

    markdown_content = re.sub(r"<!-- image -->", replace_placeholder, markdown_content, flags=re.MULTILINE)
    return markdown_content

# Função para processar arquivos XLSX (removendo fórmulas)
def process_xlsx(file_path):
    """Lê todas as planilhas de um arquivo XLSX, remove fórmulas e salva um novo arquivo com valores fixos."""
    workbook = openpyxl.load_workbook(file_path, data_only=True)  # Lê os valores calculados
    new_workbook = openpyxl.Workbook()  # Novo arquivo sem fórmulas
    new_workbook.remove(new_workbook.active)  # Remove a planilha padrão
    
    for sheet in workbook.sheetnames:
        old_sheet = workbook[sheet]
        new_sheet = new_workbook.create_sheet(title=sheet)
        
        for row in old_sheet.iter_rows(values_only=True):  # Captura apenas os valores, sem fórmulas
            new_sheet.append(row)
    
    # Salvar um novo arquivo temporário para ser enviado ao Docling
    temp_xlsx_path = os.path.join(UPLOAD_FOLDER, "temp_processed.xlsx")
    new_workbook.save(temp_xlsx_path)

    return temp_xlsx_path  # Retorna o caminho do novo arquivo

# Inicializa o banco de dados
init_db()

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# Página de login
def login_page():
    st.title("Login")
    st.write("Por favor, entre com suas credenciais para acessar o sistema.")

    username = st.text_input("Usuário", key="login_username")
    password = st.text_input("Senha", type="password", key="login_password")

    if st.button("Entrar"):
        if authenticate(username, password):
            st.session_state.logged_in = True
            st.experimental_set_query_params(logged_in="true")
            st.rerun()
        else:
            st.error("Usuário ou senha incorretos.")

# Página principal
def main_app():
    st.title("Conversor de Arquivos para Markdown")
    st.write("Faça o upload de um arquivo e ele será convertido para Markdown. Suporta PDF, DOCX e XLSX.")

    uploaded_file = st.file_uploader("Selecione um arquivo", type=["pdf", "docx", "xlsx"])

    if "progress" not in st.session_state:
        st.session_state.progress = 0

    progress_bar = st.progress(st.session_state.progress)
    if st.session_state.progress == 100:
        st.success("Processamento concluído!")

    if uploaded_file:
        if st.button("Converter"):
            try:
                st.session_state.progress = 0
                file_path = os.path.join(UPLOAD_FOLDER, uploaded_file.name)

                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                if uploaded_file.type == "application/pdf":
                    markdown_content = process_pdf_with_images_and_text(file_path, IMAGES_FOLDER)
                elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":  # DOCX
                    markdown_content = converter.convert(file_path).document.export_to_markdown()
                elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":  # XLSX
                    temp_xlsx_path = process_xlsx(file_path)  # Remove fórmulas e retorna novo XLSX
                    markdown_content = converter.convert(temp_xlsx_path).document.export_to_markdown()
                else:
                    st.error("Tipo de arquivo não suportado.")
                    return

                st.text_area("Conteúdo Convertido:", markdown_content, height=300)
                st.download_button(
                    label="Baixar como Markdown",
                    data=markdown_content,
                    file_name=f"{os.path.splitext(uploaded_file.name)[0]}.md",
                    mime="text/markdown",
                )

                st.success("Arquivo convertido com sucesso!")
            except Exception as e:
                st.error(f"Erro ao processar o arquivo: {str(e)}")

    if st.button("Sair"):
        st.session_state.logged_in = False
        st.experimental_set_query_params(logged_in="false")
        st.rerun()

if st.session_state.logged_in:
    main_app()
else:
    login_page()
