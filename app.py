import streamlit as st
import sqlite3
from bcrypt import hashpw, gensalt, checkpw
import os
import re
import openpyxl
from docling.document_converter import DocumentConverter, PdfFormatOption, WordFormatOption, ExcelFormatOption
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from imagepdf.imagepdf2 import extract_images_from_pdf, get_image_description

# Configuração de pastas
UPLOAD_FOLDER = "uploads"
CONVERTED_FOLDER = "converted"
IMAGES_FOLDER = "images"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CONVERTED_FOLDER, exist_ok=True)
os.makedirs(IMAGES_FOLDER, exist_ok=True)
STREAMLIT_RUN_ON_SAVE=False

# Configuração do banco de dados
DB_NAME = "users.db"

def init_db():
    """Inicializa o banco de dados com a tabela de usuários."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

def add_user(username: str, password: str):
    """Adiciona um novo usuário ao banco de dados."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    hashed_password = hashpw(password.encode(), gensalt())
    try:
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hashed_password))
        conn.commit()
        conn.close()
        return True, f"Usuário '{username}' criado com sucesso!"
    except sqlite3.IntegrityError:
        conn.close()
        return False, f"O usuário '{username}' já existe."
    except Exception as e:
        conn.close()
        return False, f"Erro ao criar usuário: {str(e)}"

def authenticate(username: str, password: str) -> bool:
    """Autentica o usuário comparando com as credenciais no banco de dados."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT password FROM users WHERE username = ?", (username,))
    result = cursor.fetchone()
    conn.close()
    if result and checkpw(password.encode(), result[0]):
        return True
    return False

def validate_registration_data(username: str, password: str, confirm_password: str):
    """Valida os dados de registro do usuário."""
    errors = []

    # Validar username
    if not username or len(username.strip()) == 0:
        errors.append("Nome de usuário não pode estar vazio.")
    elif len(username.strip()) < 3:
        errors.append("Nome de usuário deve ter pelo menos 3 caracteres.")
    elif len(username.strip()) > 50:
        errors.append("Nome de usuário deve ter no máximo 50 caracteres.")

    # Validar password
    if not password:
        errors.append("Senha não pode estar vazia.")
    elif len(password) < 6:
        errors.append("Senha deve ter pelo menos 6 caracteres.")
    elif len(password) > 128:
        errors.append("Senha deve ter no máximo 128 caracteres.")

    # Validar confirmação de senha
    if password != confirm_password:
        errors.append("Senhas não coincidem.")

    return errors

# Configuração do Docling para PDF, DOCX e XLSX
def create_converter():
    pdf_pipeline_options = PdfPipelineOptions(
        do_ocr=True,
        do_table_structure=True,
        extract_images=True,
        images_output_path=IMAGES_FOLDER,
        table_structure_options={"mode": "accurate"}
    )
    pdf_format_option = PdfFormatOption(pipeline_options=pdf_pipeline_options)

    word_format_option = WordFormatOption()
    excel_format_option = ExcelFormatOption()

    return DocumentConverter(
        format_options={
            InputFormat.PDF: pdf_format_option,
            InputFormat.DOCX: word_format_option,
            InputFormat.XLSX: excel_format_option
        }
    )

converter = create_converter()

# Função para processar arquivos PDF (com imagens)
def process_pdf_with_images_and_text(pdf_path, output_folder):
    """Processa o PDF para incluir texto, tabelas e imagens com descrições logo após cada <!-- image -->"""
    image_paths = extract_images_from_pdf(pdf_path, output_folder)
    
    # Extrai o conteúdo textual do PDF
    result = converter.convert(pdf_path)
    markdown_content = result.document.export_to_markdown()

    # Substituir `<!-- image -->` pela descrição correspondente
    image_index = 0
    def replace_placeholder(match):
        """Substitui <!-- image --> pela imagem e descrição correspondente."""
        nonlocal image_index
        if image_index < len(image_paths):
            image_path = image_paths[image_index]
            description = get_image_description(image_path)
            markdown_image = f"![{description}]({image_path})\n"
            image_index += 1
            return f"<!-- image -->\n{markdown_image}"
        return match.group(0)

    markdown_content = re.sub(r"<!-- image -->", replace_placeholder, markdown_content, flags=re.MULTILINE)
    return markdown_content

# Função para processar arquivos XLSX (removendo fórmulas)
def process_xlsx(file_path):
    """Lê todas as planilhas de um arquivo XLSX, remove fórmulas e salva um novo arquivo com valores fixos."""
    workbook = openpyxl.load_workbook(file_path, data_only=True)  # Lê os valores calculados
    new_workbook = openpyxl.Workbook()  # Novo arquivo sem fórmulas
    new_workbook.remove(new_workbook.active)  # Remove a planilha padrão

    for sheet in workbook.sheetnames:
        old_sheet = workbook[sheet]
        new_sheet = new_workbook.create_sheet(title=sheet)

        for row in old_sheet.iter_rows(values_only=True):  # Captura apenas os valores, sem fórmulas
            new_sheet.append(row)

    # Salvar um novo arquivo temporário para ser enviado ao Docling
    temp_xlsx_path = os.path.join(UPLOAD_FOLDER, "temp_processed.xlsx")
    new_workbook.save(temp_xlsx_path)

    return temp_xlsx_path  # Retorna o caminho do novo arquivo

# Função para processar arquivos DOCX (sem extração de imagens)
def process_standard_file(file_path):
    """Processa arquivos DOCX normalmente, sem extração de imagens."""
    result = converter.convert(file_path)
    return result.document.export_to_markdown()

# Inicializa o banco de dados
init_db()

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# Página de login e registro
def login_page():
    st.title("🔐 Acesso ao Sistema")
    st.write("Faça login ou crie uma nova conta para acessar o conversor de documentos.")

    # Criar abas para Login e Registro
    tab1, tab2 = st.tabs(["🔑 Login", "📝 Criar Conta"])

    with tab1:
        st.subheader("Entrar no Sistema")

        with st.form("login_form"):
            username = st.text_input("👤 Usuário", key="login_username", placeholder="Digite seu nome de usuário")
            password = st.text_input("🔒 Senha", type="password", key="login_password", placeholder="Digite sua senha")

            login_button = st.form_submit_button("🚀 Entrar", use_container_width=True)

            if login_button:
                if not username or not password:
                    st.error("⚠️ Por favor, preencha todos os campos.")
                elif authenticate(username, password):
                    st.session_state.logged_in = True
                    st.query_params.logged_in = "true"
                    st.success("✅ Login realizado com sucesso!")
                    st.rerun()
                else:
                    st.error("❌ Usuário ou senha incorretos.")

    with tab2:
        st.subheader("Criar Nova Conta")
        st.info("📋 Preencha os dados abaixo para criar sua conta no sistema.")

        with st.form("registration_form"):
            new_username = st.text_input("👤 Nome de Usuário", key="reg_username",
                                       placeholder="Mínimo 3 caracteres",
                                       help="Escolha um nome único para sua conta")
            new_password = st.text_input("🔒 Senha", type="password", key="reg_password",
                                       placeholder="Mínimo 6 caracteres",
                                       help="Crie uma senha segura")
            confirm_password = st.text_input("🔒 Confirmar Senha", type="password", key="reg_confirm_password",
                                           placeholder="Digite a senha novamente",
                                           help="Repita a senha para confirmação")

            register_button = st.form_submit_button("✨ Criar Conta", use_container_width=True)

            if register_button:
                # Validar dados de entrada
                validation_errors = validate_registration_data(new_username, new_password, confirm_password)

                if validation_errors:
                    for error in validation_errors:
                        st.error(f"❌ {error}")
                else:
                    # Tentar criar o usuário
                    success, message = add_user(new_username.strip(), new_password)

                    if success:
                        st.success(f"✅ {message}")
                        st.info("🎉 Conta criada com sucesso! Agora você pode fazer login na aba 'Login'.")
                        st.balloons()  # Efeito visual de celebração
                    else:
                        st.error(f"❌ {message}")

        # Informações adicionais sobre segurança
        with st.expander("🛡️ Informações de Segurança"):
            st.write("""
            **Requisitos de Senha:**
            - Mínimo de 6 caracteres
            - Máximo de 128 caracteres
            - Sua senha será criptografada com segurança

            **Requisitos de Nome de Usuário:**
            - Mínimo de 3 caracteres
            - Máximo de 50 caracteres
            - Deve ser único no sistema
            """)

# Página principal
def main_app():
    st.title("Conversor de Arquivos para Markdown")
    st.write("Faça o upload de um arquivo e ele será convertido para Markdown. Suporta PDF, DOCX e XLSX.")

    uploaded_file = st.file_uploader("Selecione um arquivo", type=["pdf", "docx", "xlsx"])

    if "progress" not in st.session_state:
        st.session_state.progress = 0

    progress_bar = st.progress(st.session_state.progress)
    if st.session_state.progress == 100:
        st.success("Processamento concluído!")

    if uploaded_file:
        if st.button("Converter"):
            try:
                st.session_state.progress = 0
                file_path = os.path.join(UPLOAD_FOLDER, uploaded_file.name)

                # Salvar o arquivo temporariamente
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                # Processar de acordo com o tipo do arquivo
                if uploaded_file.type == "application/pdf":
                    markdown_content = process_pdf_with_images_and_text(file_path, IMAGES_FOLDER)
                elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":  # DOCX
                    markdown_content = process_standard_file(file_path)
                elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":  # XLSX
                    temp_xlsx_path = process_xlsx(file_path)  # Remove fórmulas e retorna novo XLSX
                    markdown_content = converter.convert(temp_xlsx_path).document.export_to_markdown()
                else:
                    st.error("Tipo de arquivo não suportado.")
                    return

                # Exibir e baixar o resultado
                st.text_area("Conteúdo Convertido:", markdown_content, height=300)
                st.download_button(
                    label="Baixar como Markdown",
                    data=markdown_content,
                    file_name=f"{os.path.splitext(uploaded_file.name)[0]}.md",
                    mime="text/markdown",
                )

                st.success("Arquivo convertido com sucesso!")
            except Exception as e:
                st.error(f"Erro ao processar o arquivo: {str(e)}")

    if st.button("Sair"):
        st.session_state.logged_in = False
        st.query_params.logged_in = "false"
        st.rerun()

# Controle de navegação
if st.session_state.logged_in:
    main_app()
else:
    login_page()
